from pathlib import Path
import sys

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from courselens.config import COLUMN_MAP
from courselens.data_loading import inspect_workbook, load_course_data


def make_workbook(path):
    row = {col: 4 for col in COLUMN_MAP}
    row.update({"课程名称": "课程A", "选课课号": "(2023-2024-1)-TEST001-000001-1", "成绩": 90, "学分": 3, "绩点": 4.5})
    df = pd.DataFrame([row, {**row, "课程名称": "课程B", "成绩": 92, "绩点": 4.8}])
    df.to_excel(path, sheet_name="课程成绩", index=False)
    wb = load_workbook(path)
    wb.create_sheet("汇总")
    ws = wb["课程成绩"]
    ws["J2"] = "=G2+H2+I2"
    wb.save(path)


def test_loads_expected_columns_and_numeric_values(tmp_path):
    path = tmp_path / "courses.xlsx"
    make_workbook(path)
    df = load_course_data(path)
    assert len(df) == 2
    assert df["grade"].notna().all()
    assert df["grade_point"].notna().all()
    assert "offline_attendance" in df.columns


def test_detects_formula_cells(tmp_path):
    path = tmp_path / "courses.xlsx"
    make_workbook(path)
    info = inspect_workbook(path)
    assert len(info["formula_cells"]["课程成绩"]) == 1
