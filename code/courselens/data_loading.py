from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from .config import COURSE_SHEET, COLUMN_MAP, REQUIRED_COLUMNS, NUMERIC_COLUMNS


def load_course_data(path: str | Path) -> pd.DataFrame:
    path = Path(path)
    df = pd.read_excel(path, sheet_name=COURSE_SHEET, engine="openpyxl")
    df = df.rename(columns=COLUMN_MAP)
    validate_expected_columns(df)
    for col in NUMERIC_COLUMNS:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def validate_expected_columns(df: pd.DataFrame) -> None:
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing expected columns: {missing}")


def inspect_workbook(path: str | Path) -> dict:
    path = Path(path)
    wb = load_workbook(path, data_only=False)
    info = {"path": str(path), "sheets": wb.sheetnames, "formula_cells": {}}
    if COURSE_SHEET in wb.sheetnames:
        ws = wb[COURSE_SHEET]
        formulas = []
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    header = headers[cell.column - 1] if cell.column <= len(headers) else None
                    formulas.append({"cell": cell.coordinate, "column": header, "formula": cell.value})
        info["formula_cells"][COURSE_SHEET] = formulas
        info["max_row"] = ws.max_row
        info["max_column"] = ws.max_column
    return info


def schema_summary(df: pd.DataFrame, workbook_info: dict | None = None) -> dict:
    numeric = df.select_dtypes(include="number")
    return {
        "rows": int(len(df)),
        "columns": int(df.shape[1]),
        "column_names": list(df.columns),
        "dtypes": {c: str(t) for c, t in df.dtypes.items()},
        "missing": {c: int(v) for c, v in df.isna().sum().items()},
        "numeric_ranges": {
            c: {"min": float(numeric[c].min()), "max": float(numeric[c].max())}
            for c in numeric.columns if numeric[c].notna().any()
        },
        "workbook": workbook_info or {},
    }
